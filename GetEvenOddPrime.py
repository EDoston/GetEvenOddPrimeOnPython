import math
import os
from openpyxl import Workbook

with open("list_of_random_numbers.txt", 'r') as txt:
    for line in txt:
        data = line.split(', ')

randomNumbers = [int(x) for x in data]
even = [num for num in randomNumbers if num % 2 == 0]
odd = [num for num in randomNumbers if num % 2 != 0]

prime = []

def find_prime_numbers(n):
    for x in randomNumbers:
        if is_prime(x):
            prime.append(x)

def is_prime(n):
    if n == 1:
        return False
        
    for i in range(2, int(math.sqrt(n)+1)):
        if (n % i) == 0:
            return False
    return True

find_prime_numbers(randomNumbers);

# Exclude prime from odd numbers
odd = [o for o in odd if o not in prime]

# Delete file if exit 
if os.path.exists('Result.xlsx'):
    os.remove('Result.xlsx')

wb = Workbook();
ws = wb.active

def write_even(work_sheet, even):
    work_sheet.cell(row=1, column = 1).value = 'Even numbers'
    work_sheet.column_dimensions['A'].width = 14;

    i = 1
    for e in even:
        i += 1
        work_sheet.cell(row = i, column = 1).value = e

def write_odd(work_sheet, odd):
    work_sheet.cell(row = 1, column = 2).value = 'Odd numbers'
    work_sheet.column_dimensions['B'].width = 14

    i = 1
    for o in odd:
        i += 1
        work_sheet.cell(row = i, column = 2).value = o

def write_prime(work_sheet, prime):
    work_sheet.cell(row = 1, column = 3).value = 'Prime numbers'
    work_sheet.column_dimensions['C'].width = 14

    i = 1
    for p in prime:
        i += 1
        work_sheet.cell(row = i, column = 3).value = p

write_even(ws, even)
write_odd(ws, odd)
write_prime(ws, prime)

wb.save('Result.xlsx');